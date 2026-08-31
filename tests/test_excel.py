import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook, load_workbook

from tools import siteexcel


class SiteExcelTest(unittest.TestCase):
    def test_roundtrip_preserves_styles_dimensions_and_other_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'sites.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = '站点'
            ws.append(['公益站', '状态'])
            ws.append(['示例站', '有效'])
            ws['A2'].font = ws['A2'].font.copy(bold=True, color='FFFF0000')
            ws['B2'].fill = ws['B2'].fill.copy(fgColor='FFFFFF00', fill_type='solid')
            ws['A2'].font = ws['A2'].font.copy(italic=True)
            ws['B2'].alignment = ws['B2'].alignment.copy(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[2].height = 32
            ws.column_dimensions['A'].width = 28
            wb.create_sheet('说明')['A1'] = '保留内容'
            wb.save(path)

            document = siteexcel.load_workbook(path)
            self.assertEqual(document['rows'][0]['name'], '示例站')
            self.assertEqual(document['columns'][0]['width'], 28)
            self.assertEqual(document['row_heights'][2], 32)
            name_style = next(style for key, style in document['styles'].items() if key.endswith('|name'))
            self.assertEqual(name_style['font']['bold'], True)
            self.assertEqual(name_style['font']['color'], '#FF0000')
            status_style = next(style for key, style in document['styles'].items() if key.endswith('|status'))
            self.assertEqual(status_style['fillColor'], '#FFFF00')
            self.assertEqual(status_style['horizontal'], 'center')
            self.assertEqual(status_style['vertical'], 'center')
            self.assertEqual(status_style['wrapText'], True)
            document['rows'][0]['status'] = '失效'
            siteexcel.save_workbook(path, document)

            reopened = load_workbook(path, data_only=False)
            self.assertEqual(reopened['站点']['B2'].value, '失效')
            self.assertTrue(reopened['站点']['A2'].font.bold)
            self.assertTrue(reopened['站点']['A2'].font.italic)
            self.assertEqual(reopened['站点']['B2'].alignment.horizontal, 'center')
            self.assertTrue(reopened['站点']['B2'].alignment.wrap_text)
            self.assertEqual(reopened['说明']['A1'].value, '保留内容')

    def test_corrupted_workbook_has_clear_error(self):
        with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
            tmp.write(b'not xlsx')
            tmp.flush()
            with self.assertRaisesRegex(ValueError, 'XLSX'):
                siteexcel.load_workbook(Path(tmp.name))


if __name__ == '__main__':
    unittest.main()

