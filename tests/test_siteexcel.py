import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

import sys
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / 'tools'))
import siteexcel


class SiteExcelMergeTest(unittest.TestCase):
    def test_load_and_save_preserve_merges_and_other_sheets(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / 'sites.xlsx'
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = '站点'
            worksheet.append(['名称', '状态', 'uid'])
            worksheet.append(['测试站', '正常', 'row-1'])
            worksheet.merge_cells('B2:C2')
            other = workbook.create_sheet('其他')
            other['A1'] = '保留'
            workbook.save(path)

            document = siteexcel.load_workbook(path, [{'field': 'name', 'header': '名称'}, {'field': 'status', 'header': '状态'}])
            self.assertEqual(document['merges'], ['B2:C2'])
            document['merges'] = ['A2:B2']
            siteexcel.save_workbook(path, document)

            saved = load_workbook(path)
            self.assertEqual([str(item) for item in saved.worksheets[0].merged_cells.ranges], ['A2:B2'])
            self.assertEqual(saved['其他']['A1'].value, '保留')
            self.assertEqual(saved.worksheets[0]['A2'].value, '测试站')


if __name__ == '__main__':
    unittest.main()
