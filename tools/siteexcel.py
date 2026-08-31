# -*- coding: utf-8 -*-
"""XLSX 业务数据层：第一张工作表承载站点数据，其余工作表原样保留。"""
import copy
import json
import os
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook as openpyxl_load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SITES_XLSX = ROOT / "public" / "sites.xlsx"
SITES_CSV = ROOT / "public" / "sites.csv"
COLUMNS_JSON = ROOT / "public" / "columns.json"
BUTTONS_JSON = ROOT / "public" / "buttons.json"

FIELDS = [
    ("name", "公益站"), ("status", "状态"), ("rating", "评分"), ("register", "注册"),
    ("daily", "每日签到"), ("invite", "邀请制"), ("model", "模型质量"), ("exp", "体验感"),
    ("other", "其他"), ("other2", "其他2·白嫖org"), ("other3", "其他3·飞书合集"),
    ("other4", "其他4·幻城导航"), ("verified", "验证"), ("models", "模型"),
    ("latency", "响应"), ("api_status", "渠道状态"), ("url", "注册链接"), ("checkin", "签到地址"),
]
HIDDEN_FIELD, HIDDEN_HEADER = "hidden", "隐藏"
FIELD_BY_HEADER = {header: field for field, header in FIELDS} | {HIDDEN_HEADER: HIDDEN_FIELD, "uid": "uid"}


def _clean(value):
    return "" if value is None else str(value).strip()


def _color_hex(color):
    if not color:
        return None
    value = color.rgb if color.type == "rgb" else None
    if value and len(value) == 8:
        return "#" + value[2:].upper()
    return None


def _style_from_cell(cell):
    style = {
        "font": {
            "bold": bool(cell.font.bold) or None,
            "italic": bool(cell.font.italic) or None,
            "color": _color_hex(cell.font.color),
        },
        "fillColor": _color_hex(cell.fill.fgColor) if cell.fill.fill_type == "solid" else None,
        "horizontal": cell.alignment.horizontal,
        "vertical": cell.alignment.vertical,
        "wrapText": bool(cell.alignment.wrap_text) or None,
    }
    if not any((style["font"]["bold"], style["font"]["italic"], style["font"]["color"], style["fillColor"], style["horizontal"], style["vertical"], style["wrapText"])):
        return None
    return style


def _column_def(header, meta_by_header, width):
    metadata = dict(meta_by_header.get(header, {}))
    metadata["header"] = header
    metadata["field"] = metadata.get("field") or FIELD_BY_HEADER.get(header, header)
    if width is not None:
        metadata["width"] = width
    return metadata


def load_workbook(path=SITES_XLSX, columns_meta=None):
    """读取第一张工作表，返回行、列、样式、行高、列宽等编辑器数据。"""
    path = Path(path)
    if not path.exists():
        raise ValueError(f"XLSX 文件不存在: {path}")
    try:
        workbook = openpyxl_load_workbook(path)
    except Exception as exc:
        raise ValueError(f"XLSX 工作簿读取失败: {exc}") from exc
    if not workbook.worksheets:
        raise ValueError("XLSX 工作簿缺少第一张工作表")
    worksheet = workbook.worksheets[0]
    meta_by_header = {item.get("header"): item for item in (columns_meta or []) if item.get("header")}
    headers = [_clean(cell.value) for cell in worksheet[1] if _clean(cell.value)]
    columns = []
    for index, header in enumerate(headers, start=1):
        definition = _column_def(header, meta_by_header, worksheet.column_dimensions[get_column_letter(index)].width)
        if definition["field"] not in (HIDDEN_FIELD, "uid"):
            columns.append(definition)
    all_columns = [_column_def(header, meta_by_header, worksheet.column_dimensions[get_column_letter(index)].width) for index, header in enumerate(headers, start=1)]
    rows, styles, row_heights, column_widths = [], {}, {}, {}
    for index, column in enumerate(all_columns, start=1):
        width = worksheet.column_dimensions[get_column_letter(index)].width
        if width is not None:
            column_widths[get_column_letter(index)] = width
    for row_number, excel_row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        item = {column["field"]: _clean(excel_row[index - 1].value) for index, column in enumerate(all_columns, start=1)}
        if not any(item.values()):
            continue
        item.setdefault("uid", "row-" + str(row_number - 1))
        rows.append(item)
        if worksheet.row_dimensions[row_number].height is not None:
            row_heights[row_number] = worksheet.row_dimensions[row_number].height
        for index, column in enumerate(all_columns, start=1):
            style = _style_from_cell(excel_row[index - 1])
            if style:
                styles[item["uid"] + "|" + column["field"]] = style
    return {
        "rows": rows,
        "columns": columns,
        "styles": styles,
        "row_heights": row_heights,
        "column_widths": column_widths,
        "worksheet_name": worksheet.title,
        "sheet_names": workbook.sheetnames,
    }


def _apply_style(cell, style):
    font = style.get("font") or {}
    if any(value is not None for value in (font.get("bold"), font.get("italic"), font.get("color"))):
        cell.font = copy.copy(cell.font)
        cell.font = Font(name=cell.font.name, sz=cell.font.sz, b=font.get("bold", cell.font.bold), i=font.get("italic", cell.font.italic), color=("FF" + font["color"].replace("#", "")) if font.get("color") else cell.font.color)
    if style.get("fillColor"):
        cell.fill = PatternFill(fill_type="solid", fgColor="FF" + style["fillColor"].replace("#", ""))
    if any(style.get(key) is not None for key in ("horizontal", "vertical", "wrapText")):
        cell.alignment = Alignment(horizontal=style.get("horizontal") or cell.alignment.horizontal, vertical=style.get("vertical") or cell.alignment.vertical, wrap_text=style.get("wrapText") if style.get("wrapText") is not None else cell.alignment.wrap_text)


def save_workbook(path, document):
    """把编辑器数据写回第一张工作表，使用临时文件原子替换并保留其他工作表。"""
    path = Path(path)
    if not path.exists():
        workbook = Workbook()
        workbook.active.title = "站点"
    else:
        try:
            workbook = openpyxl_load_workbook(path)
        except Exception as exc:
            raise ValueError(f"XLSX 工作簿读取失败: {exc}") from exc
    worksheet = workbook.worksheets[0]
    if worksheet.max_row:
        worksheet.delete_rows(1, worksheet.max_row)
    rows = document.get("rows") or []
    columns = document.get("columns") or []
    has_hidden = any(HIDDEN_FIELD in row for row in rows)
    has_uid = any(row.get("uid") for row in rows)
    headers = ([HIDDEN_HEADER] if has_hidden else []) + [column.get("header") or column.get("field") for column in columns] + (["uid"] if has_uid else [])
    worksheet.append(headers)
    fields = ([HIDDEN_FIELD] if has_hidden else []) + [column.get("field") for column in columns] + (["uid"] if has_uid else [])
    styles = document.get("styles") or {}
    row_heights = {int(key): value for key, value in (document.get("row_heights") or document.get("rowHeights") or {}).items()}
    column_widths = document.get("column_widths") or document.get("columnWidths") or {}
    for row_index, item in enumerate(rows, start=2):
        worksheet.append([item.get(field) or "" for field in fields])
        if row_index in row_heights:
            worksheet.row_dimensions[row_index].height = row_heights[row_index]
        for column_index, column in enumerate(columns, start=1 + int(has_hidden)):
            style = styles.get(str(item.get("uid", "")) + "|" + str(column.get("field")))
            if style:
                _apply_style(worksheet.cell(row_index, column_index), style)
    for letter, width in column_widths.items():
        worksheet.column_dimensions[str(letter)].width = width
    worksheet.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(prefix=path.stem + ".", suffix=".tmp.xlsx", dir=path.parent, delete=False) as tmp:
        temp_path = Path(tmp.name)
    try:
        workbook.save(temp_path)
        os.replace(temp_path, path)
    finally:
        if temp_path.exists():
            temp_path.unlink()
    return {"ok": True, "path": str(path), "rows": len(rows)}


def migrate_csv(csv_path=SITES_CSV, xlsx_path=SITES_XLSX):
    """首次将旧 CSV 数据迁移为 XLSX，并保留 CSV 备份。"""
    import csv
    with open(csv_path, encoding="utf-8-sig", newline="") as handle:
        source_rows = list(csv.DictReader(handle))
    meta = json.loads(COLUMNS_JSON.read_text(encoding="utf-8")) if COLUMNS_JSON.exists() else {"columns": []}
    meta_by_header = {item.get("header"): item for item in meta.get("columns", [])}
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "站点"
    headers = [HIDDEN_HEADER] + [header for _field, header in FIELDS] + ["uid"]
    worksheet.append(headers)
    for index, raw in enumerate(source_rows, start=1):
        row = {FIELD_BY_HEADER.get(key.strip(), key.strip()): _clean(value) for key, value in raw.items() if key}
        row.setdefault("uid", "row-" + str(index))
        worksheet.append([row.get(field, "") for field, _header in [(HIDDEN_FIELD, HIDDEN_HEADER)] + FIELDS + [("uid", "uid")]])
    for index, (_field, header) in enumerate([(HIDDEN_FIELD, HIDDEN_HEADER)] + FIELDS + [("uid", "uid")], start=1):
        definition = meta_by_header.get(header, {})
        if definition.get("width"):
            worksheet.column_dimensions[get_column_letter(index)].width = definition["width"]
    worksheet.freeze_panes = "A2"
    workbook.save(xlsx_path)
    return len(source_rows)


def _load_columns_meta():
    if not COLUMNS_JSON.exists():
        return []
    payload = json.loads(COLUMNS_JSON.read_text(encoding='utf-8'))
    return payload.get('columns') or []


def load_rows(path=SITES_XLSX):
    return load_workbook(path, _load_columns_meta()).get('rows', [])


def load_meta():
    if not COLUMNS_JSON.exists():
        return {'columns': []}
    return json.loads(COLUMNS_JSON.read_text(encoding='utf-8'))


def save_meta(meta):
    COLUMNS_JSON.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')


def touch_updated(meta=None, day=None):
    from datetime import date
    payload = meta or load_meta()
    payload['updated'] = (day or date.today().isoformat())
    save_meta(payload)
    return payload


def _normalise_columns(coldefs, current):
    if not coldefs:
        return current
    current_by_field = {item.get('field'): item for item in current}
    columns = []
    for item in coldefs:
        if isinstance(item, dict):
            field = item.get('field')
            header = item.get('header') or field
            value = dict(current_by_field.get(field, {}))
            value.update(item)
        else:
            field, header = item
            value = dict(current_by_field.get(field, {}))
            value.update(field=field, header=header)
        if value.get('field') not in (None, HIDDEN_FIELD, 'uid'):
            columns.append(value)
    return columns


def save_rows(rows, coldefs=None, path=SITES_XLSX):
    current = load_workbook(path, _load_columns_meta()) if Path(path).exists() else {
        'rows': [], 'columns': [], 'styles': {}, 'row_heights': {}, 'column_widths': {}
    }
    document = {
        'rows': rows,
        'columns': _normalise_columns(coldefs, current.get('columns') or []),
        'styles': current.get('styles') or {},
        'row_heights': current.get('row_heights') or {},
        'column_widths': current.get('column_widths') or {},
    }
    return save_workbook(path, document)
